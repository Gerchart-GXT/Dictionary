import openpyxl

class HandleExcel:
    def __init__(self, fileName):
        self.fileName = fileName
        self.__workBook = openpyxl.load_workbook(fileName)
        self.sheetList = self.__workBook.sheetnames
    def _read_sheet(self, sheetName):
        sheet = self.__workBook[sheetName]
        result = []
        for row in sheet.iter_rows():
            val = []
            for cell in row:
                val.append(cell.value)
            result.append(val)
        return result
    def _write_sheet(self, sheetName, valList):
        sheet = None
        try:
            sheet = self.__workBook[sheetName]
        except KeyError as e:
            print(sheetName + "不存在， 将为您创建")
            sheet = self.__workBook.create_sheet(title=sheetName)
        lastRow = 1
        while sheet[lastRow][0].value != None:
            lastRow += 1
        for column, value in enumerate(valList, start = 1):
            sheet.cell(row = lastRow, column = column, value = value)
    def _clear_sheet(self, sheetName):
        sheet = self.__workBook[sheetName]
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None
    def _delete_sheet(self, sheetName):
        sheet_to_delete = self.__workBook[sheetName]
        self.__workBook.remove(sheet_to_delete)
    def _save(self):
        self.__workBook.save(self.fileName)

def main():
    workBook = None
    try:
        workBook = HandleExcel("test.xlsx")
        workBook._read_sheet('Sheet1')
        valList = [
            ['Data1', 'Data2', 'Data3'],
            ['Data4', 'Data5', 'Data6'],
            ['Data7', 'Data8', 'Data9']
        ]
        workBook._write_sheet("Sheet2",valList)
        workBook._save()
    except Exception as e:
        print(e)
        exit(1)
        
if __name__ == "__main__":
    main()